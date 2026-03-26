"""
Eurocontrol FRA (Free Route Airspace) Points source.

Downloads and parses the FRA Points list from Eurocontrol, which contains
~26,000+ named waypoints used in European free route airspace. Updated
every AIRAC cycle (28 days).

Publication page: https://www.eurocontrol.int/publication/free-route-airspace-fra-points-list-ecac-area
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import requests

from ..models.euro_aip_model import EuroAipModel
from ..models.waypoint import Waypoint
from ..utils.dms_parser import parse_fra_latitude, parse_fra_longitude
from .base import SourceInterface
from .cached import CachedSource

logger = logging.getLogger(__name__)

PUBLICATION_URL = "https://www.eurocontrol.int/publication/free-route-airspace-fra-points-list-ecac-area"

# NAVAID type codes from the Excel "Point Type" column
_NAVAID_TYPES = {"VOR", "DME", "VORDME", "VORTAC", "NDB", "LOCATOR", "NDBDME"}


class EurocontrolFRASource(CachedSource, SourceInterface):
    """Source for Eurocontrol FRA waypoint data.

    Downloads the latest FRA Points Excel file from Eurocontrol,
    parses waypoint names, coordinates, and metadata, and adds
    them to the model as Waypoint objects.
    """

    def __init__(self, cache_dir: str, local_path: Optional[str] = None):
        """
        Args:
            cache_dir: Directory for caching downloaded files.
            local_path: Optional path to a locally downloaded Excel file.
                        If provided, skips the download and uses this file.
        """
        super().__init__(cache_dir)
        self.local_path = Path(local_path) if local_path else None

    def update_model(self, model: EuroAipModel, airports: Optional[List[str]] = None) -> None:
        """Update the model with FRA waypoints."""
        waypoints = self.get_waypoints()
        result = model.bulk_add_waypoints(waypoints)
        logger.info(
            "EurocontrolFRA: added %d, updated %d waypoints",
            result["added"], result["updated"],
        )

    def get_waypoints(self, max_age_days: int = 28) -> List[Waypoint]:
        """Get FRA waypoints, using cache if available.

        Args:
            max_age_days: Cache validity in days (default 28 = 1 AIRAC cycle)

        Returns:
            List of Waypoint objects
        """
        excel_data = self._get_excel_data(max_age_days)
        return self._parse_excel(excel_data)

    def _get_excel_data(self, max_age_days: int) -> bytes:
        """Get the Excel file data, from local path, cache, or download."""
        if self.local_path and self.local_path.exists():
            logger.info("Using local FRA file: %s", self.local_path)
            return self.local_path.read_bytes()

        # Check cache
        cache_file = self._get_cache_file("fra_points", "xlsx")
        is_valid, reason = self._is_cache_valid(cache_file, max_age_days)
        if is_valid:
            logger.info("Using cached FRA file: %s", cache_file)
            return cache_file.read_bytes()

        # Download
        logger.info("Downloading FRA points from Eurocontrol (cache %s)", reason)
        data = self._download_latest()
        self._save_to_cache(data, "fra_points", "xlsx")
        return data

    def _download_latest(self) -> bytes:
        """Scrape the publication page and download the latest Excel file."""
        url = self._find_latest_download_url()
        logger.info("Downloading FRA Excel from: %s", url)
        response = requests.get(url, timeout=60)
        response.raise_for_status()
        return response.content

    def _find_latest_download_url(self) -> str:
        """Scrape the Eurocontrol publication page for the latest .xlsx URL."""
        response = requests.get(PUBLICATION_URL, timeout=30)
        response.raise_for_status()

        # Find all .xlsx download links
        pattern = r'href="(/sites/default/files/[^"]*\.xlsx)"'
        matches = re.findall(pattern, response.text)

        if not matches:
            raise RuntimeError(
                "Could not find any .xlsx download links on the Eurocontrol FRA page"
            )

        # The first match is typically the most recent file
        latest_path = matches[0]
        return f"https://www.eurocontrol.int{latest_path}"

    def _parse_excel(self, data: bytes) -> List[Waypoint]:
        """Parse the FRA Points Excel file into Waypoint objects."""
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        # Find the data sheet
        sheet_name = "FRA Points"
        if sheet_name not in wb.sheetnames:
            # Fallback: try first non-cover sheet
            for name in wb.sheetnames:
                if name.upper() != "COVER" and "EXPLANATION" not in name.upper():
                    sheet_name = name
                    break
            else:
                raise RuntimeError(f"Could not find data sheet in FRA Excel. Sheets: {wb.sheetnames}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Parse header row
        header = [str(h).strip() if h else "" for h in rows[0]]
        col_idx = {name: i for i, name in enumerate(header)}

        # Required columns
        required = {"FRA Point", "FRA Point Latitude", "FRA Point Longitude"}
        missing = required - set(col_idx.keys())
        if missing:
            raise RuntimeError(f"Missing required columns in FRA Excel: {missing}")

        waypoints: Dict[str, Waypoint] = {}
        skipped = 0

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                wp = self._parse_row(row, col_idx, row_num)
                if wp is None:
                    skipped += 1
                    continue

                if wp.name in waypoints:
                    # Merge FIR codes for duplicate names
                    existing = waypoints[wp.name]
                    if wp.fir_codes:
                        existing_firs = set(existing.fir_list)
                        new_firs = set(wp.fir_list)
                        merged = sorted(existing_firs | new_firs)
                        existing.fir_codes = ",".join(merged)
                else:
                    waypoints[wp.name] = wp
            except Exception as e:
                logger.debug("Row %d: parse error: %s", row_num, e)
                skipped += 1

        wb.close()
        logger.info(
            "Parsed %d unique waypoints from FRA Excel (%d rows skipped)",
            len(waypoints), skipped,
        )
        return list(waypoints.values())

    def _parse_row(self, row: tuple, col_idx: Dict[str, int], row_num: int) -> Optional[Waypoint]:
        """Parse a single Excel row into a Waypoint, or None to skip."""

        def get(col_name: str) -> Optional[str]:
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        # Skip deleted rows
        change_record = get("Change Record")
        if change_record and change_record.upper() == "DEL":
            return None

        name = get("FRA Point")
        lat_str = get("FRA Point Latitude")
        lon_str = get("FRA Point Longitude")

        if not name or not lat_str or not lon_str:
            return None

        name = name.strip().upper()
        if not name:
            return None

        # Parse coordinates
        try:
            lat = parse_fra_latitude(lat_str)
            lon = parse_fra_longitude(lon_str)
        except ValueError as e:
            logger.debug("Row %d (%s): coordinate parse error: %s", row_num, name, e)
            return None

        # Determine point type
        point_type_raw = get("Point Type")
        if point_type_raw and point_type_raw.upper() in _NAVAID_TYPES:
            point_type = point_type_raw.upper()
        else:
            point_type = "5LNC"

        # FIR codes from "Loc. indicators" column
        fir_raw = get("Loc. indicators")
        fir_codes = None
        if fir_raw:
            # May be comma-separated
            firs = [f.strip() for f in fir_raw.split(",") if f.strip()]
            if firs:
                fir_codes = ",".join(sorted(firs))

        level_availability = get("Level Availability")

        return Waypoint(
            name=name,
            latitude_deg=lat,
            longitude_deg=lon,
            point_type=point_type,
            fir_codes=fir_codes,
            level_availability=level_availability,
            source="eurocontrol_fra",
        )

    def _save_to_cache(self, data, key: str, ext: str) -> None:
        """Override to support xlsx binary files (base class only handles json/csv/pdf)."""
        if ext == "xlsx":
            cache_file = self._get_cache_file(key, ext)
            with open(cache_file, "wb") as f:
                f.write(data)
        else:
            super()._save_to_cache(data, key, ext)

    def _load_from_cache(self, key: str, ext: str):
        """Override to support loading xlsx binary files."""
        if ext == "xlsx":
            cache_file = self._get_cache_file(key, ext)
            return cache_file.read_bytes()
        return super()._load_from_cache(key, ext)
